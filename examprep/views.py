import csv
import html
import re

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import redirect_to_login
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.db.models import F, Count, Q, Max
from django.http import Http404, HttpResponse
from django.utils.html import strip_tags
from django.utils.translation import gettext as _
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import (ExamTrack, Topic, Lesson, LessonBlock, LessonProgress,
                     WritingDrill, WritingDrillProgress, WRITING_DRILL_POINTS,
                     LESSON_POINTS, SKILL_CHOICES, SKILL_ICONS,
                     GrammarPoint, GrammarSynonym,
                     GRAMMAR_CATEGORY_CHOICES, GRAMMAR_FUNCTION_CHOICES,
                     VocabRoot, VocabEntry, VocabRelation,
                     VOCAB_TOPIC_CHOICES, VOCAB_POS_CHOICES)
from prime.subjects import get_study_subjects, value_visible


def _strip_tags_plain(value):
    """HTML note → plain text for the spreadsheet export.

    strip_tags alone leaves entities behind (&nbsp;, &rarr;), which look like
    noise in a cell, so unescape after stripping and collapse the whitespace
    the removed block tags leave.
    """
    text = html.unescape(strip_tags(value or ''))
    return re.sub(r'[ \t]*\n\s*', '\n', re.sub(r'[ \t]+', ' ', text)).strip()


def _can_edit(user, lesson):
    """Staff, or the lesson's own author, may edit on-page."""
    if not user.is_authenticated:
        return False
    return user.is_staff or (lesson.author_id and lesson.author_id == user.id)


def _published_filter(user):
    """Staff see drafts too; everyone else only published items."""
    return {} if user.is_staff else {'is_published': True}


def examprep_home(request):
    """List published exam tracks as cards, each with its skill chips."""
    tracks = list(
        ExamTrack.objects
        .filter(is_published=True)
        .annotate(lesson_count=Count('lessons', filter=Q(lessons__is_published=True)))
    )
    # Study-subject preference: applies unless the visitor asked for ?all=1
    personalized = False
    slugs = get_study_subjects(request)
    if slugs and not request.GET.get('all'):
        tracks = [t for t in tracks
                  if value_visible(t.slug, slugs, 'examprep_tracks')]
        personalized = True
    skill_labels = dict(SKILL_CHOICES)
    for track in tracks:
        skills = (track.lessons.filter(is_published=True)
                  .values_list('skill', flat=True).distinct())
        track.skill_chips = [
            {'value': value, 'label': skill_labels[value],
             'icon': SKILL_ICONS.get(value, 'bi-journal-text')}
            for value, _ in SKILL_CHOICES if value in skills
        ]
    return render(request, 'examprep/home.html', {
        'tracks': tracks,
        'personalized': personalized,
    })


def track_detail(request, track_slug):
    """One track; its skills (Reading, Writing, ...) shown as cards."""
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)

    lessons = list(track.lessons.filter(**pub))
    topics = list(track.topics.filter(**pub))

    # Group by skill, preserving SKILL_CHOICES order.
    groups = []
    for value, label in SKILL_CHOICES:
        skill_lessons = [l for l in lessons if l.skill == value]
        if skill_lessons:
            groups.append({
                'value':       value,
                'label':       label,
                'icon':        SKILL_ICONS.get(value, 'bi-journal-text'),
                'count':       len(skill_lessons),
                'topic_count': sum(1 for t in topics if t.skill == value),
            })

    return render(request, 'examprep/track_detail.html', {
        'track':         track,
        'groups':        groups,
        'drill_count':   track.writing_drills.filter(**pub).count(),
        'grammar_count': track.grammar_points.filter(**pub).count(),
        'vocab_count':   track.vocab_entries.filter(**pub).count(),
        'root_count':    track.vocab_roots.filter(**pub).count(),
    })


def skill_detail(request, track_slug, skill):
    """One skill inside a track: its question-type topics as cards, each card
    listing that topic's lessons. Lessons without a topic form a final group."""
    skill_label = dict(SKILL_CHOICES).get(skill)
    if skill_label is None:
        raise Http404('Unknown section.')

    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)

    lessons = list(track.lessons.filter(skill=skill, **pub))
    if not lessons:
        raise Http404('No lessons in this section yet.')

    groups = []
    for topic in track.topics.filter(skill=skill, **pub):
        topic_lessons = [l for l in lessons if l.topic_id == topic.id]
        if topic_lessons:
            groups.append({'topic': topic, 'lessons': topic_lessons})

    grouped_ids = {l.id for g in groups for l in g['lessons']}
    loose_lessons = [l for l in lessons if l.id not in grouped_ids]

    return render(request, 'examprep/skill_detail.html', {
        'track':         track,
        'skill':         skill,
        'skill_label':   skill_label,
        'skill_icon':    SKILL_ICONS.get(skill, 'bi-journal-text'),
        'groups':        groups,
        'loose_lessons': loose_lessons,
        'lesson_count':  len(lessons),
    })


def lesson_detail(request, track_slug, skill, slug):
    """A lesson 'player': its blocks, plus prev/next + a jump list for the skill."""
    pub = _published_filter(request.user)
    lesson = get_object_or_404(
        Lesson.objects.select_related('track', 'topic'),
        track__slug=track_slug, skill=skill, slug=slug, **pub,
    )

    # Ordered siblings drive the playlist: lessons of the same topic when the
    # lesson belongs to one, otherwise the skill's remaining ungrouped lessons.
    siblings_qs = lesson.track.lessons.filter(skill=skill, **pub)
    if lesson.topic_id:
        siblings_qs = siblings_qs.filter(topic_id=lesson.topic_id)
    else:
        siblings_qs = siblings_qs.filter(topic__isnull=True)
    siblings = list(siblings_qs)
    index = siblings.index(lesson)
    prev_lesson = siblings[index - 1] if index > 0 else None
    next_lesson = siblings[index + 1] if index < len(siblings) - 1 else None

    blocks = list(lesson.blocks.prefetch_related('choices').all())
    has_question = any(b.choices.exists() for b in blocks)

    # Older lessons open with an <h2> repeating the lesson title, which stacks
    # a third copy under the breadcrumb and page <h1>. Hide that one heading at
    # render time (display only — nothing is saved).
    if blocks and blocks[0].rich_text:
        m = re.match(r'\s*<h2[^>]*>(.*?)</h2>', blocks[0].rich_text, re.S | re.I)
        if m:
            heading = re.sub(r'<[^>]+>', '', m.group(1))
            norm = lambda s: re.sub(r'\s+', ' ', s).strip().casefold()
            if norm(heading) == norm(lesson.title):
                blocks[0].rich_text = blocks[0].rich_text[m.end():]

    submitted = request.method == 'POST'
    if submitted:
        # Stateless check: mark each choice and the block result, no DB writes.
        for block in blocks:
            choices = list(block.choices.all())
            if not choices:
                continue
            raw = request.POST.get(f'mcq_{block.id}')
            selected_id = int(raw) if (raw and raw.isdigit()) else None
            correct = next((c for c in choices if c.is_correct), None)
            block.selected_id = selected_id
            block.is_correct = bool(correct and selected_id == correct.id)
            block.answered = selected_id is not None
            for choice in choices:
                choice.was_selected = (choice.id == selected_id)
    else:
        # Count a view only on plain reads, not on answer submissions.
        Lesson.objects.filter(pk=lesson.pk).update(views=F('views') + 1)

    is_finished = (
        request.user.is_authenticated
        and LessonProgress.objects.filter(user=request.user, lesson=lesson).exists()
    )

    skill_label = dict(SKILL_CHOICES).get(skill, skill)
    return render(request, 'examprep/lesson_detail.html', {
        'is_finished':  is_finished,
        'lesson_points': LESSON_POINTS,
        'lesson':       lesson,
        'topic':        lesson.topic,
        'skill':        skill,
        'skill_label':  skill_label,
        'skill_icon':   SKILL_ICONS.get(skill, 'bi-journal-text'),
        'blocks':       blocks,
        'siblings':     siblings,
        'current_no':   index + 1,
        'total_no':     len(siblings),
        'prev_lesson':  prev_lesson,
        'next_lesson':  next_lesson,
        'has_question': has_question,
        'submitted':    submitted,
        'can_edit':     _can_edit(request.user, lesson),
    })


@login_required
def lesson_edit(request, track_slug, skill, slug):
    """On-page editor for a lesson and its content blocks (author/staff only)."""
    lesson = get_object_or_404(
        Lesson.objects.select_related('track'),
        track__slug=track_slug, skill=skill, slug=slug,
    )
    if not _can_edit(request.user, lesson):
        raise PermissionDenied

    blocks = list(lesson.blocks.all())

    if request.method == 'POST':
        lesson.title = (request.POST.get('title') or lesson.title).strip()
        lesson.summary = (request.POST.get('summary') or '')[:300]
        lesson.is_published = bool(request.POST.get('is_published'))
        try:
            lesson.order = int(request.POST.get('order', lesson.order))
        except (TypeError, ValueError):
            pass
        lesson.save()

        for b in blocks:
            if request.POST.get(f'delete_{b.id}'):
                b.delete()
                continue
            b.rich_text = (request.POST.get(f'rich_text_{b.id}') or '') or None
            b.explanation = (request.POST.get(f'explanation_{b.id}') or '') or None
            b.caption = (request.POST.get(f'caption_{b.id}') or '')[:300]
            try:
                b.order = int(request.POST.get(f'order_{b.id}', b.order))
            except (TypeError, ValueError):
                pass
            b.save()

        new_html = (request.POST.get('new_rich_text') or '').strip()
        if new_html:
            nxt = (lesson.blocks.aggregate(m=Max('order'))['m'] or 0) + 1
            LessonBlock.objects.create(lesson=lesson, order=nxt, rich_text=new_html)

        messages.success(request, 'Saqlandi / Saved.')
        return redirect('examprep_lesson', track_slug=lesson.track.slug,
                        skill=skill, slug=lesson.slug)

    return render(request, 'examprep/lesson_edit.html', {
        'lesson': lesson,
        'skill':  skill,
        'blocks': blocks,
    })


@login_required
@require_POST
def lesson_finish(request, track_slug, skill, slug):
    """Mark a lesson finished; award points to the panda once."""
    lesson = get_object_or_404(
        Lesson.objects.select_related('track'),
        track__slug=track_slug, skill=skill, slug=slug,
        **_published_filter(request.user),
    )
    progress, created = LessonProgress.objects.get_or_create(
        user=request.user, lesson=lesson,
        defaults={'points_awarded': LESSON_POINTS},
    )
    if created:
        try:
            request.user.profile.panda.recalc_rating()
            messages.success(request, _('Lesson finished! +%(points)d points')
                             % {'points': LESSON_POINTS})
        except Exception:
            messages.success(request, _('Lesson finished!'))
    else:
        messages.info(request, _('You already finished this lesson.'))
    return redirect('examprep_lesson', track_slug=track_slug, skill=skill, slug=slug)


# ── Writing drills ─────────────────────────────────────────────────────────

def drill_list(request, track_slug):
    """A track's interactive writing drills, grouped by exam question type."""
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)
    drills = list(track.writing_drills.filter(**pub))

    finished_ids = set()
    if request.user.is_authenticated:
        finished_ids = set(WritingDrillProgress.objects
                           .filter(user=request.user, drill__track=track)
                           .values_list('drill_id', flat=True))

    # Group in QTYPE_CHOICES order, keeping only the types this track uses —
    # so IELTS never shows TOPIK's 51-54 numbering, or vice versa.
    groups = {}
    for drill in drills:
        drill.is_finished = drill.id in finished_ids
        groups.setdefault(drill.get_qtype_display(), []).append(drill)

    return render(request, 'examprep/drill_list.html', {
        'track': track,
        'groups': groups,
        'finished_count': len(finished_ids),
        'total_count': len(drills),
    })


def drill_detail(request, track_slug, pk):
    """One drill: exam question + chart, key expressions, fill-in scaffold,
    model answer behind a reveal, flashcards, finish button."""
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)
    drill = get_object_or_404(
        WritingDrill.objects.select_related('track'), pk=pk, track=track, **pub,
    )

    siblings = list(track.writing_drills.filter(qtype=drill.qtype, **pub))
    index = siblings.index(drill)
    prev_d = siblings[index - 1] if index > 0 else None
    next_d = siblings[index + 1] if index < len(siblings) - 1 else None

    is_finished = (request.user.is_authenticated and
                   WritingDrillProgress.objects
                   .filter(user=request.user, drill=drill).exists())

    WritingDrill.objects.filter(pk=drill.pk).update(views=F('views') + 1)

    return render(request, 'examprep/drill_detail.html', {
        'track': track,
        'drill': drill,
        'words': list(drill.words.all()),
        'current_no': index + 1,
        'total_no': len(siblings),
        'prev_d': prev_d,
        'next_d': next_d,
        'is_finished': is_finished,
        'drill_points': WRITING_DRILL_POINTS,
    })


@login_required
@require_POST
def drill_finish(request, track_slug, pk):
    """Mark a drill finished; award points to the panda once."""
    drill = get_object_or_404(
        WritingDrill, pk=pk, track__slug=track_slug, **_published_filter(request.user),
    )
    progress, created = WritingDrillProgress.objects.get_or_create(
        user=request.user, drill=drill,
        defaults={'points_awarded': WRITING_DRILL_POINTS},
    )
    if created:
        try:
            request.user.profile.panda.recalc_rating()
            messages.success(request, _('Drill finished! +%(points)d points')
                             % {'points': WRITING_DRILL_POINTS})
        except Exception:
            messages.success(request, _('Drill finished!'))
    else:
        messages.info(request, _('You already finished this drill.'))
    return redirect('examprep_drill', track_slug=track_slug, pk=pk)


# ── Grammar bank ───────────────────────────────────────────────────────────
# One filterable table of every grammar pattern in a track, plus a print sheet
# and a spreadsheet export. Filtering is entirely server-side (GET params) so
# the page works with JavaScript off, which is also what makes the print and
# download views trivial: they read the same filters and render the same rows.

def _qs_with(request, **overrides):
    """Current query string with some params replaced — the href for one chip.

    Built here rather than in the template because Django templates cannot
    manipulate a QueryDict; a value of None drops the param (that is how a chip
    toggles itself off).
    """
    params = request.GET.copy()
    for key, value in overrides.items():
        if value is None:
            params.pop(key, None)
        else:
            params[key] = value
    encoded = params.urlencode()
    return f'?{encoded}' if encoded else '?'


def _grammar_filters(request, track):
    """Read the filter params, and build the chip lists the template renders.

    Returns (queryset, context) — shared by the list, print and download views
    so all three always agree on what "the current selection" means.
    """
    pub = _published_filter(request.user)
    qs = (GrammarPoint.objects.filter(track=track, **pub)
          .prefetch_related('examples', 'synonyms__related'))

    q = (request.GET.get('q') or '').strip()
    level = (request.GET.get('level') or '').strip()
    category = (request.GET.get('cat') or '').strip()
    function = (request.GET.get('fn') or '').strip()
    group_by = 'function' if request.GET.get('by') == 'function' else 'category'

    if q:
        qs = qs.filter(
            Q(pattern__icontains=q) | Q(meaning__icontains=q) |
            Q(attach__icontains=q) | Q(note__icontains=q) |
            Q(examples__korean__icontains=q) | Q(examples__uz__icontains=q) |
            Q(synonyms__pattern__icontains=q)
        ).distinct()
    if level.isdigit():
        qs = qs.filter(level=int(level))
    if category in dict(GRAMMAR_CATEGORY_CHOICES):
        qs = qs.filter(category=category)
    if function in dict(GRAMMAR_FUNCTION_CHOICES):
        qs = qs.filter(function=function)

    points = list(qs)

    # Chip counts come from the unfiltered set, so a chip never reads "0" just
    # because another filter is active — the student can always see what else
    # is in the bank and switch to it in one tap.
    all_points = list(GrammarPoint.objects.filter(track=track, **pub)
                      .values_list('category', 'function', 'level'))
    cat_counts, fn_counts, lvl_counts = {}, {}, {}
    for c, f, l in all_points:
        cat_counts[c] = cat_counts.get(c, 0) + 1
        fn_counts[f] = fn_counts.get(f, 0) + 1
        lvl_counts[l] = lvl_counts.get(l, 0) + 1

    context = {
        'track':        track,
        'points':       points,
        'total':        len(all_points),
        'q':            q,
        'level':        level,
        'category':     category,
        'function':     function,
        'group_by':     group_by,
        'has_filter':   bool(q or level or category or function),
        'categories':   [{'value': c, 'label': l, 'count': cat_counts.get(c, 0),
                          'on': category == c,
                          'url': _qs_with(request, cat=None if category == c else c)}
                         for c, l in GRAMMAR_CATEGORY_CHOICES if cat_counts.get(c)],
        'functions':    [{'value': f, 'label': l, 'count': fn_counts.get(f, 0),
                          'on': function == f,
                          'url': _qs_with(request, fn=None if function == f else f)}
                         for f, l in GRAMMAR_FUNCTION_CHOICES if fn_counts.get(f)],
        'levels':       [{'value': n, 'count': lvl_counts.get(n, 0),
                          'on': level == str(n),
                          'url': _qs_with(request, level=None if level == str(n) else n)}
                         for n in range(1, 7) if lvl_counts.get(n)],
        'url_clear':     _qs_with(request, q=None, level=None, cat=None, fn=None),
        'url_by_cat':    _qs_with(request, by=None),
        'url_by_fn':     _qs_with(request, by='function'),
        'querystring':   request.GET.urlencode(),
    }
    return points, context


def _grammar_groups(points, group_by):
    """Split the rows into the sections the page shows, in choices order."""
    choices = (GRAMMAR_FUNCTION_CHOICES if group_by == 'function'
               else GRAMMAR_CATEGORY_CHOICES)
    buckets = {}
    for p in points:
        buckets.setdefault(getattr(p, group_by), []).append(p)
    return [{'value': value, 'label': label, 'points': buckets[value]}
            for value, label in choices if value in buckets]


def grammar_list(request, track_slug):
    """The grammar summary table: filter chips + grouped rows, each expandable."""
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)
    points, context = _grammar_filters(request, track)
    if not context['total']:
        raise Http404('No grammar entries in this track yet.')
    context['groups'] = _grammar_groups(points, context['group_by'])
    return render(request, 'examprep/grammar_list.html', context)


def grammar_detail(request, track_slug, slug):
    """One pattern in full: every example, the nuance notes, and its synonym
    set — plus the other patterns that share its meaning group."""
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)
    point = get_object_or_404(
        GrammarPoint.objects.select_related('track')
        .prefetch_related('examples', 'synonyms__related'),
        track=track, slug=slug, **pub,
    )

    # "Shu ma'nodagi boshqalar" — siblings in the same meaning group. This is
    # the whole reason `function` exists: comparing -아서 with -니까 is the
    # question students actually have.
    siblings = list(GrammarPoint.objects
                    .filter(track=track, function=point.function, **pub)
                    .exclude(pk=point.pk)[:12])

    # Patterns that name THIS one as their synonym, but that it does not name
    # back — otherwise the comparison only works in one direction.
    named_ids = {s.related_id for s in point.synonyms.all() if s.related_id}
    incoming = [s for s in GrammarSynonym.objects
                .filter(related=point, point__track=track, **{f'point__{k}': v
                                                              for k, v in pub.items()})
                .select_related('point')
                if s.point_id not in named_ids]

    return render(request, 'examprep/grammar_detail.html', {
        'track':    track,
        'point':    point,
        'siblings': siblings,
        'incoming': incoming,
    })


def _require_staff(request):
    """Gate for the take-away formats (print sheet, spreadsheet export).

    Reading the bank on-site is open to everyone; walking away with a
    distributable copy of it is not. Anonymous visitors are sent to log in
    (they may simply not be signed in yet); a signed-in non-staff user gets a
    403, because for them it is a permission answer, not a login prompt.
    """
    if not request.user.is_authenticated:
        return redirect_to_login(request.get_full_path())
    if not request.user.is_staff:
        raise PermissionDenied
    return None


def grammar_print(request, track_slug):
    """A dense, ink-friendly sheet of the current selection — Ctrl+P → PDF.

    Staff only — see _require_staff.
    """
    denied = _require_staff(request)
    if denied:
        return denied
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)
    points, context = _grammar_filters(request, track)
    context['groups'] = _grammar_groups(points, context['group_by'])
    context['compact'] = request.GET.get('compact') == '1'
    context['url_compact'] = _qs_with(request, compact='1')
    context['url_full'] = _qs_with(request, compact=None)
    # The download link in the print bar must not carry `compact` through.
    context['querystring'] = _qs_with(request, compact=None).lstrip('?')
    return render(request, 'examprep/grammar_print.html', context)


def grammar_download(request, track_slug):
    """Export the current selection as .xlsx (default) or .csv. Staff only."""
    denied = _require_staff(request)
    if denied:
        return denied
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)
    points, _ctx = _grammar_filters(request, track)

    header = ['Pattern', 'TOPIK', 'Turi', 'Ma\'nosi (guruh)', 'Ma\'nosi',
              'Qo\'shilishi', 'Shakl qoidasi', 'Namuna (한국어)', 'Tarjima',
              'Sinonimlar', 'Izoh', 'Ko\'p uchraydi']

    def row_of(p):
        examples = list(p.examples.all())
        synonyms = '; '.join(
            f'{s.pattern} — {s.note}' if s.note else s.pattern
            for s in p.synonyms.all()
        )
        return [
            p.pattern,
            p.level,
            p.get_category_display(),
            p.get_function_display(),
            p.meaning,
            p.attach,
            _strip_tags_plain(p.form_rule),
            '\n'.join(e.korean for e in examples),
            '\n'.join(e.uz for e in examples),
            synonyms,
            _strip_tags_plain(p.note),
            p.stars,
        ]

    fmt = request.GET.get('fmt', 'xlsx')
    stem = f'{track.slug}-grammatika'

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{stem}.csv"'
        # BOM so Excel opens the Hangul correctly instead of showing mojibake.
        response.write('﻿')
        writer = csv.writer(response)
        writer.writerow(header)
        for p in points:
            writer.writerow(row_of(p))
        return response

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Grammatika'
    sheet.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4F46E5')
        cell.alignment = Alignment(vertical='center')
    for p in points:
        sheet.append(row_of(p))
    for cell_row in sheet.iter_rows(min_row=2):
        for cell in cell_row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for column, width in zip('ABCDEFGHIJKL',
                             [16, 7, 22, 24, 30, 24, 32, 40, 40, 40, 40, 8]):
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{stem}.xlsx"'
    workbook.save(response)
    return response


# ── Vocabulary bank ────────────────────────────────────────────────────────
# Mirrors the grammar bank (open to read, staff-only to take away) and adds a
# root-family view: the same words regrouped under the Sino-Korean morpheme
# they share, which is where most of the learning leverage is.

def _vocab_filters(request, track):
    """Read the filter params and build the chip lists. Same contract as
    _grammar_filters, so the list, roots, print and download views all agree
    on what "the current selection" is."""
    pub = _published_filter(request.user)
    qs = (VocabEntry.objects.filter(track=track, **pub)
          .prefetch_related('examples', 'relations__related', 'roots'))

    q = (request.GET.get('q') or '').strip()
    level = (request.GET.get('level') or '').strip()
    topic = (request.GET.get('topic') or '').strip()
    pos = (request.GET.get('pos') or '').strip()
    root = (request.GET.get('root') or '').strip()

    if q:
        qs = qs.filter(
            Q(word__icontains=q) | Q(meaning__icontains=q) | Q(hanja__icontains=q) |
            Q(collocation__icontains=q) | Q(note__icontains=q) |
            Q(examples__korean__icontains=q) | Q(examples__uz__icontains=q) |
            Q(roots__syllable__icontains=q) | Q(relations__word__icontains=q)
        ).distinct()
    if level.isdigit():
        qs = qs.filter(level=int(level))
    if topic in dict(VOCAB_TOPIC_CHOICES):
        qs = qs.filter(topic=topic)
    if pos in dict(VOCAB_POS_CHOICES):
        qs = qs.filter(pos=pos)
    if root:
        qs = qs.filter(roots__slug=root).distinct()

    entries = list(qs)

    # Chip counts come from the unfiltered set so a chip never reads "0" just
    # because another filter is active.
    all_entries = list(VocabEntry.objects.filter(track=track, **pub)
                       .values_list('topic', 'pos', 'level'))
    topic_counts, pos_counts, lvl_counts = {}, {}, {}
    for t, p, l in all_entries:
        topic_counts[t] = topic_counts.get(t, 0) + 1
        pos_counts[p] = pos_counts.get(p, 0) + 1
        lvl_counts[l] = lvl_counts.get(l, 0) + 1

    active_root = None
    if root:
        active_root = VocabRoot.objects.filter(track=track, slug=root, **pub).first()

    context = {
        'track':       track,
        'entries':     entries,
        'total':       len(all_entries),
        'q':           q,
        'level':       level,
        'topic':       topic,
        'pos':         pos,
        'root':        root,
        'active_root': active_root,
        'has_filter':  bool(q or level or topic or pos or root),
        'topics':      [{'value': t, 'label': l, 'count': topic_counts.get(t, 0),
                         'on': topic == t,
                         'url': _qs_with(request, topic=None if topic == t else t)}
                        for t, l in VOCAB_TOPIC_CHOICES if topic_counts.get(t)],
        'poses':       [{'value': p, 'label': l, 'count': pos_counts.get(p, 0),
                         'on': pos == p,
                         'url': _qs_with(request, pos=None if pos == p else p)}
                        for p, l in VOCAB_POS_CHOICES if pos_counts.get(p)],
        'levels':      [{'value': n, 'count': lvl_counts.get(n, 0),
                         'on': level == str(n),
                         'url': _qs_with(request, level=None if level == str(n) else n)}
                        for n in range(1, 7) if lvl_counts.get(n)],
        'url_clear':   _qs_with(request, q=None, level=None, topic=None, pos=None, root=None),
        'querystring': request.GET.urlencode(),
    }
    return entries, context


def _vocab_groups(entries, active_root=None):
    """Split rows into topic sections, in VOCAB_TOPIC_CHOICES order.

    When a root filter is active the topics are the wrong axis: a family of
    eight words scatters into six one-row sections. Show it as a single table
    named after the root instead — that is what the reader asked for.
    """
    if active_root is not None:
        return [{'value': active_root.slug,
                 'label': f'{active_root.label} — {active_root.meaning}',
                 'entries': entries}] if entries else []
    buckets = {}
    for e in entries:
        buckets.setdefault(e.topic, []).append(e)
    return [{'value': value, 'label': label, 'entries': buckets[value]}
            for value, label in VOCAB_TOPIC_CHOICES if value in buckets]


def vocab_list(request, track_slug):
    """The vocabulary table: filter chips + rows grouped by theme."""
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)
    entries, context = _vocab_filters(request, track)
    if not context['total']:
        raise Http404('No vocabulary in this track yet.')
    context['groups'] = _vocab_groups(entries, context['active_root'])
    context['root_count'] = track.vocab_roots.filter(**pub).count()
    return render(request, 'examprep/vocab_list.html', context)


def vocab_roots(request, track_slug):
    """Word families: every root with the words built on it.

    The point of the page — 출(出) once, then 출구·출근·출발·출석·제출·수출
    read as one family instead of six unrelated words.
    """
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)

    roots = list(track.vocab_roots.filter(**pub).prefetch_related('entries'))
    if not roots:
        raise Http404('No word roots in this track yet.')

    q = (request.GET.get('q') or '').strip()
    if q:
        roots = [r for r in roots
                 if q in r.syllable or q in r.hanja or q.casefold() in r.meaning.casefold()
                 or any(q in e.word for e in r.entries.all())]

    families = []
    for r in roots:
        words = [e for e in r.entries.all() if e.is_published or request.user.is_staff]
        if words:
            families.append({'root': r, 'words': sorted(words, key=lambda e: (e.level, e.order))})

    return render(request, 'examprep/vocab_roots.html', {
        'track':       track,
        'families':    families,
        'q':           q,
        'total_roots': len(families),
        'total_words': sum(len(f['words']) for f in families),
    })


def vocab_detail(request, track_slug, slug):
    """One word in full: examples, collocations, its roots and the other words
    built on them, plus synonyms and antonyms."""
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)
    entry = get_object_or_404(
        VocabEntry.objects.select_related('track')
        .prefetch_related('examples', 'relations__related', 'roots__entries'),
        track=track, slug=slug, **pub,
    )

    # Siblings from every root this word is built on — the family view, scoped.
    seen, family = {entry.id}, []
    for root in entry.roots.all():
        words = [e for e in root.entries.all()
                 if e.id not in seen and (e.is_published or request.user.is_staff)]
        for e in words:
            seen.add(e.id)
        if words:
            family.append({'root': root, 'words': words})

    # Words that name THIS one as a relation without it naming them back.
    named = {r.related_id for r in entry.relations.all() if r.related_id}
    incoming = [r for r in VocabRelation.objects
                .filter(related=entry, entry__track=track)
                .select_related('entry')
                if r.entry_id not in named and (r.entry.is_published or request.user.is_staff)]

    return render(request, 'examprep/vocab_detail.html', {
        'track':    track,
        'entry':    entry,
        'family':   family,
        'incoming': incoming,
        'same_topic': list(VocabEntry.objects
                           .filter(track=track, topic=entry.topic, **pub)
                           .exclude(pk=entry.pk)[:12]),
    })


def vocab_print(request, track_slug):
    """Watermarked, ink-friendly sheet of the current selection. Staff only."""
    denied = _require_staff(request)
    if denied:
        return denied
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)
    entries, context = _vocab_filters(request, track)

    # ?by=root prints the word families instead of the thematic table — the
    # study sheet you actually want to stick on a wall.
    by_root = request.GET.get('by') == 'root'
    if by_root:
        ids = {e.id for e in entries}
        families = []
        for r in track.vocab_roots.filter(**pub).prefetch_related('entries'):
            words = [e for e in r.entries.all() if e.id in ids]
            if words:
                families.append({'root': r,
                                 'words': sorted(words, key=lambda e: (e.level, e.order))})
        context['families'] = families
    else:
        context['groups'] = _vocab_groups(entries, context['active_root'])

    context['by_root'] = by_root
    context['compact'] = request.GET.get('compact') == '1'
    context['url_compact'] = _qs_with(request, compact='1')
    context['url_full'] = _qs_with(request, compact=None)
    context['url_by_root'] = _qs_with(request, by='root')
    context['url_by_topic'] = _qs_with(request, by=None)
    context['querystring'] = _qs_with(request, compact=None).lstrip('?')
    return render(request, 'examprep/vocab_print.html', context)


def vocab_download(request, track_slug):
    """Export the current selection as .xlsx (default) or .csv. Staff only."""
    denied = _require_staff(request)
    if denied:
        return denied
    pub = _published_filter(request.user)
    track = get_object_or_404(ExamTrack, slug=track_slug, **pub)
    entries, _ctx = _vocab_filters(request, track)

    header = ['So\'z', 'Hanja', 'TOPIK', 'So\'z turkumi', 'Mavzu', 'Ma\'nosi',
              'O\'zaklar', 'Birikmalar', 'Namuna (한국어)', 'Tarjima',
              'Sinonim', 'Antonim', 'Izoh', 'Ko\'p uchraydi']

    def row_of(e):
        examples = list(e.examples.all())
        rel = lambda kind: '; '.join(
            f'{r.word} — {r.note}' if r.note else r.word
            for r in e.relations.all() if r.kind == kind
        )
        return [
            e.word, e.hanja, e.level, e.get_pos_display(), e.get_topic_display(),
            e.meaning,
            ' · '.join(r.label for r in e.roots.all()),
            e.collocation,
            '\n'.join(x.korean for x in examples),
            '\n'.join(x.uz for x in examples),
            rel('syn'), rel('ant'),
            _strip_tags_plain(e.note),
            e.stars,
        ]

    fmt = request.GET.get('fmt', 'xlsx')
    stem = f'{track.slug}-lugat'

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{stem}.csv"'
        response.write('﻿')   # BOM so Excel reads the Hangul correctly
        writer = csv.writer(response)
        writer.writerow(header)
        for e in entries:
            writer.writerow(row_of(e))
        return response

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Lug\'at'
    sheet.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='0F766E')
        cell.alignment = Alignment(vertical='center')
    for e in entries:
        sheet.append(row_of(e))
    for cell_row in sheet.iter_rows(min_row=2):
        for cell in cell_row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for column, width in zip('ABCDEFGHIJKLMN',
                             [16, 10, 7, 16, 24, 34, 16, 28, 38, 38, 26, 26, 34, 8]):
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{stem}.xlsx"'
    workbook.save(response)
    return response
